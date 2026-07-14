"""
Contabilidad de produccion: Google Sheets (gspread + cuenta de servicio) con
respaldo local en Excel (openpyxl) cuando no hay credenciales — la fase
financiera funciona end-to-end desde el primer dia.

Configuracion Google Sheets (una sola vez):
1. En Google Cloud Console crea un proyecto, habilita la API "Google Sheets".
2. Crea una CUENTA DE SERVICIO y descarga su JSON de credenciales; guardalo en
   config/google_service_account.json (o apunta GOOGLE_SA_JSON a otra ruta).
3. Crea un spreadsheet y compartelo (Editor) con el client_email del JSON.
4. Copia el ID del spreadsheet (la parte larga de la URL) a
   SHEETS_SPREADSHEET_ID en .env.

Hojas gestionadas:
- LibroProduccion : una fila por evento de produccion (ingreso/costo/margen).
- ResumenMensual  : agregado por mes y SKU.
- PlanCompras     : plan MRP activo (lo que se envio/enviara a Odoo).
"""
from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
import sys

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from config import settings
from integrations import state_store

ENCABEZADOS_LIBRO = ["ts", "linea", "sku", "unidades", "precio_unit_cop",
                     "costo_unit_cop", "ingreso_cop", "costo_cop", "margen_cop"]


def numero_cop(texto, default: float | None = None) -> float | None:
    """Convierte una celda del libro (formato colombiano: punto = separador
    de miles, coma = separador decimal -- "3.850" -> 3850.0, "1.200,0" ->
    1200.0) a float. NO es el formato ingles (coma de miles, punto decimal).
    None/vacio/texto no numerico -> default. El '%' se tolera (se descarta;
    quien llama decide si dividir por 100 -- ver core.finanzas_negocio._num)."""
    if texto is None or texto == "":
        return default
    if isinstance(texto, (int, float)):
        return float(texto)
    limpio = str(texto).strip().replace("$", "").replace(" ", "").rstrip("%")
    limpio = limpio.replace(".", "").replace(",", ".") if "," in limpio \
        else limpio.replace(".", "")
    try:
        return float(limpio)
    except ValueError:
        return default


class Contabilidad:
    def __init__(self) -> None:
        self.modo = "sheets" if (settings.SHEETS_ENABLED and not settings.DRY_RUN_FORZADO) else "excel"
        self._ss = None
        if settings.EXTERNAL_ONLY and self.modo != "sheets":
            raise RuntimeError("EXTERNAL_ONLY=true: Google Sheets es obligatorio")

    @staticmethod
    def _sin_fallback(error: Exception) -> None:
        if settings.EXTERNAL_ONLY:
            raise RuntimeError(f"Google Sheets no disponible en modo EXTERNAL_ONLY: {error}") from error

    # ------------------------------------------------------------- backends
    def _spreadsheet(self):
        if self._ss is None:
            import gspread
            gc = gspread.service_account(filename=settings.GOOGLE_SA_JSON)
            self._ss = gc.open_by_key(settings.SHEETS_SPREADSHEET_ID)
        return self._ss

    def _hoja(self, nombre: str, encabezados: list[str]):
        ss = self._spreadsheet()
        try:
            ws = ss.worksheet(nombre)
        except Exception:  # noqa: BLE001 — WorksheetNotFound
            ws = ss.add_worksheet(nombre, rows=2000, cols=max(12, len(encabezados)))
            ws.append_row(encabezados)
        if not ws.row_values(1):
            ws.append_row(encabezados)
        return ws

    def leer_hoja(self, nombre: str, rango: str | None = None,
                   encabezado: int = 1) -> pd.DataFrame:
        """Lee una tabla de Sheets como fuente externa de verdad.

        `encabezado` es 1-based dentro del rango devuelto. En modo estricto
        una hoja ausente o vacia es un error visible, nunca un fallback local.
        """
        try:
            ws = self._spreadsheet().worksheet(nombre)
            filas = ws.get(rango, value_render_option="UNFORMATTED_VALUE") \
                if rango else ws.get(value_render_option="UNFORMATTED_VALUE")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            return pd.DataFrame()
        if len(filas) < encabezado:
            if settings.EXTERNAL_ONLY:
                raise ValueError(f"Hoja {nombre!r} vacia o sin encabezado")
            return pd.DataFrame()
        columnas = [str(c).strip() for c in filas[encabezado - 1]]
        ancho = len(columnas)
        datos = [(list(f) + [""] * ancho)[:ancho] for f in filas[encabezado:]
                 if any(str(c).strip() for c in f)]
        return pd.DataFrame(datos, columns=columnas)

    def leer_demanda(self, escenario: bool = False) -> pd.DataFrame:
        hoja = "DemandaEscenario" if escenario else "Demanda"
        from core.forecast import normalizar_demanda_mensual
        return normalizar_demanda_mensual(self.leer_hoja(hoja, "A4:F16"))

    def leer_inventarios(self) -> pd.DataFrame:
        return self.leer_hoja("Inventarios", "A4:I8")

    def leer_plan_compras(self) -> pd.DataFrame:
        return self.leer_hoja("PlanCompras", "A4:K")

    def leer_libro_produccion(self) -> pd.DataFrame:
        return self.leer_hoja("LibroProduccion")

    def leer_resumen_mensual(self) -> pd.DataFrame:
        return self.leer_hoja("ResumenMensual", "A4:F")

    def leer_kpis_uns(self) -> pd.DataFrame:
        return self.leer_hoja("KPIs_UNS", "A4:G")

    def leer_costos_unitarios(self) -> dict[str, float]:
        """Costos variables por SKU calculados en la hoja ``Costos_Lote``.

        Se localizan por la etiqueta ``COSTO UNITARIO`` dentro de cada bloque,
        no por una coordenada fija, para soportar que el usuario agregue rubros.
        """
        try:
            filas = self._spreadsheet().worksheet("Costos_Lote").get(
                value_render_option="UNFORMATTED_VALUE")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            return {}
        skus = ["P1-CC350-RGB", "P2-QT1500-PET", "P3-GARR25L"]
        out: dict[str, float] = {}
        bloque = -1
        for fila in filas:
            etiqueta = str(fila[0]).strip() if fila else ""
            if etiqueta.startswith("P1 ·"):
                bloque = 0
            elif etiqueta.startswith("P2 ·"):
                bloque = 1
            elif etiqueta.startswith("P3 ·"):
                bloque = 2
            elif etiqueta.startswith("COSTO UNITARIO") and bloque >= 0 and len(fila) > 1:
                out[skus[bloque]] = float(fila[1])
        if settings.EXTERNAL_ONLY and len(out) != len(skus):
            raise ValueError("Costos_Lote no contiene los tres COSTO UNITARIO")
        return out

    def leer_maestro_productos(self) -> pd.DataFrame:
        """Maestro fisico y comercial vivo; reemplaza maestro_productos.csv."""
        df = self.leer_hoja("Maestro_Productos")
        requeridas = {"sku", "nombre", "linea", "ean13", "litros_por_unidad",
                      "unidades_por_caja", "cajas_por_pallet",
                      "precio_venta_cop", "costo_material_cop"}
        faltan = requeridas - set(df.columns)
        if faltan:
            raise ValueError(f"Maestro_Productos sin columnas: {sorted(faltan)}")
        numericas = ["litros_por_unidad", "unidades_por_caja", "cajas_por_pallet",
                     "precio_venta_cop", "costo_material_cop",
                     "participacion_categoria"]
        for c in numericas:
            if c in df:
                df[c] = pd.to_numeric(df[c], errors="raise")
        df["ean13"] = df["ean13"].astype(str).str.replace(r"\.0$", "", regex=True)
        return df

    def leer_clientes(self) -> pd.DataFrame:
        df = self.leer_hoja("Clientes")
        requeridas = {"nombre", "ciudad", "canal", "participacion"}
        faltan = requeridas - set(df.columns)
        if faltan:
            raise ValueError(f"Clientes sin columnas: {sorted(faltan)}")
        df["participacion"] = pd.to_numeric(df["participacion"], errors="raise")
        return df

    def leer_dataset_pronostico(self, hoja: str) -> pd.DataFrame:
        permitidas = {"Forecast_Historico_Mensual", "Forecast_Historico_Trimestral",
                      "Forecast_Perfil_Formato", "Forecast_KOF_Trimestral",
                      "Forecast_Pronostico_Mensual", "Forecast_Pronostico_Trimestral",
                      "Forecast_Metricas", "Forecast_QA"}
        if hoja not in permitidas:
            raise ValueError(f"dataset de pronostico no permitido: {hoja}")
        df = self.leer_hoja(hoja)
        if hoja == "Forecast_Pronostico_Mensual":
            from core.forecast import normalizar_demanda_mensual
            df = normalizar_demanda_mensual(df)
        return df

    def leer_config_pronostico(self, clave: str) -> dict:
        df = self.leer_hoja("Forecast_Configuracion")
        if not {"clave", "json"}.issubset(df.columns):
            raise ValueError("Forecast_Configuracion requiere columnas clave/json")
        filas = df[df["clave"].astype(str) == clave]
        if filas.empty:
            raise KeyError(f"configuracion de pronostico ausente: {clave}")
        return json.loads(str(filas.iloc[0]["json"]))

    def publicar_tabla_externa(self, hoja: str, df: pd.DataFrame) -> str:
        """Reemplaza una tabla maestra externa durante migracion controlada."""
        return self._escribir(hoja, list(df.columns),
                              df.fillna("").astype(object).values.tolist(),
                              reemplazar=True)

    def _excel_append(self, hoja: str, encabezados: list[str], filas: list[list]) -> None:
        from openpyxl import Workbook, load_workbook
        path = settings.LEDGER_XLSX
        wb = load_workbook(path) if path.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and hoja != "Sheet":
            wb.remove(wb["Sheet"])
        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws.append(encabezados)
        ws = wb[hoja]
        for f in filas:
            ws.append(f)
        wb.save(path)

    def _escribir(self, hoja: str, encabezados: list[str], filas: list[list],
                  reemplazar: bool = False) -> str:
        if not filas:
            return self.modo
        if self.modo == "sheets":
            try:
                ws = self._hoja(hoja, encabezados)
                if reemplazar:
                    ws.clear()
                    ws.append_row(encabezados)
                ws.append_rows(filas, value_input_option="USER_ENTERED")
                state_store.log("sheets", f"append:{hoja}", f"{len(filas)} filas")
                return "sheets"
            except Exception as e:  # noqa: BLE001 — degradar a Excel local
                state_store.log("sheets", "ERROR -> fallback excel", str(e))
                self._sin_fallback(e)
                self.modo = "excel"
        if reemplazar and settings.LEDGER_XLSX.exists():
            # borrar y recrear la hoja EN LA MISMA sesion: el libro nunca queda
            # con cero hojas visibles (openpyxl no permite guardarlo asi)
            from openpyxl import load_workbook
            wb = load_workbook(settings.LEDGER_XLSX)
            if hoja in wb.sheetnames:
                idx = wb.sheetnames.index(hoja)
                wb.remove(wb[hoja])
                ws = wb.create_sheet(hoja, idx)
                ws.append(encabezados)
                for f in filas:
                    ws.append(f)
                wb.save(settings.LEDGER_XLSX)
                state_store.log("contabilidad", f"excel:{hoja}",
                                f"{len(filas)} filas (reemplazo)")
                return "excel"
        self._excel_append(hoja, encabezados, filas)
        state_store.log("contabilidad", f"excel:{hoja}", f"{len(filas)} filas")
        return "excel"

    # ------------------------------------------------------------- operaciones
    def registrar_produccion(self, eventos: list[dict],
                             maestro: pd.DataFrame) -> tuple[str, pd.DataFrame]:
        """eventos: [{'ts','linea','sku','qty'}...] -> libro con margen."""
        m = maestro.set_index("sku")
        filas = []
        for ev in eventos:
            if ev["sku"] not in m.index:
                continue
            p = m.loc[ev["sku"]]
            ing = ev["qty"] * p["precio_venta_cop"]
            cos = ev["qty"] * p["costo_material_cop"]
            filas.append([ev.get("ts", datetime.now(timezone.utc).isoformat(timespec="seconds")),
                          ev.get("linea", ""), ev["sku"], ev["qty"],
                          float(p["precio_venta_cop"]), float(p["costo_material_cop"]),
                          round(ing), round(cos), round(ing - cos)])
        destino = self._escribir("LibroProduccion", ENCABEZADOS_LIBRO, filas)
        return destino, pd.DataFrame(filas, columns=ENCABEZADOS_LIBRO)

    def publicar_resumen_mensual(self, resumen: pd.DataFrame) -> str:
        return self._escribir("ResumenMensual", list(resumen.columns),
                              resumen.astype(object).values.tolist(), reemplazar=True)

    def publicar_plan_compras(self, plan: pd.DataFrame) -> str:
        return self._escribir("PlanCompras", list(plan.columns),
                              plan.astype(object).values.tolist(), reemplazar=True)


    def sincronizar_libro_completo(self, eventos: list[dict],
                                   maestro: pd.DataFrame) -> tuple[str, int]:
        """Reconstruye y REEMPLAZA LibroProduccion con todos los eventos
        (idempotente: sin duplicados al sincronizar varias veces)."""
        m = maestro.set_index("sku")
        filas = []
        for ev in eventos:
            if ev["sku"] not in m.index:
                continue
            p = m.loc[ev["sku"]]
            ing = ev["qty"] * p["precio_venta_cop"]
            cos = ev["qty"] * p["costo_material_cop"]
            filas.append([ev.get("ts", ""), ev.get("linea", ""), ev["sku"], ev["qty"],
                          float(p["precio_venta_cop"]), float(p["costo_material_cop"]),
                          round(ing), round(cos), round(ing - cos)])
        destino = self._escribir("LibroProduccion", ENCABEZADOS_LIBRO, filas,
                                 reemplazar=True)
        return destino, len(filas)


    # ------------------------------------------------------ integracion ERP v4
    def publicar_tiempos(self, demanda_mensual: pd.DataFrame | None = None) -> str:
        """Publica/actualiza la hoja Tiempos (estudio de tiempos por linea)."""
        from core.tiempos_oee import tabla_tiempos
        df = tabla_tiempos(demanda_mensual).fillna("")
        return self._escribir("Tiempos", list(df.columns),
                              df.astype(object).values.tolist(), reemplazar=True)

    def publicar_oee(self) -> str:
        """Publica/actualiza la hoja OEE (base medido vs +5% justificado)."""
        from core.tiempos_oee import tabla_oee
        df = tabla_oee().fillna("")
        return self._escribir("OEE", list(df.columns),
                              df.astype(object).values.tolist(), reemplazar=True)

    def registrar_kpis_uns(self, kpis: list[dict]) -> str:
        """Anexa KPIs recibidos del UNS (ts, linea, rama, kpi, valor)."""
        enc = ["ts", "linea", "rama", "kpi", "valor_num", "valor_txt", "topic"]
        filas = [[k.get(c, "") if k.get(c) is not None else "" for c in enc]
                 for k in kpis]
        return self._escribir("KPIs_UNS", enc, filas)

    # ---- contrato de demanda: rangos FIJOS (las hojas Financiero del libro
    # ---- referencian Demanda!D5:F16 y DemandaEscenario!D5:F16; por eso NO se
    # ---- usa clear+append sino escritura posicional que preserva las formulas)
    ENC_DEMANDA = ["escenario", "mes_num", "etiqueta",
                   "P1-CC350-RGB_unidades", "P2-QT1500-PET_unidades",
                   "P3-GARR25L_unidades"]

    def _escribir_rango(self, hoja: str, filas: list[list],
                        fila_inicio: int = 4) -> str:
        """Escribe un bloque en posicion fija (A{fila_inicio}) sin tocar el
        resto de la hoja. Crea la hoja si no existe."""
        if self.modo == "sheets":
            try:
                ss = self._spreadsheet()
                try:
                    ws = ss.worksheet(hoja)
                except Exception:  # noqa: BLE001
                    ws = ss.add_worksheet(hoja, rows=60, cols=20)
                ws.update(filas, f"A{fila_inicio}",
                          value_input_option="USER_ENTERED")
                state_store.log("sheets", f"rango:{hoja}",
                                f"{len(filas)} filas @A{fila_inicio}")
                return "sheets"
            except Exception as e:  # noqa: BLE001
                state_store.log("sheets", "ERROR -> fallback excel", str(e))
                self._sin_fallback(e)
                self.modo = "excel"
        from openpyxl import Workbook, load_workbook
        path = settings.LEDGER_XLSX
        wb = load_workbook(path) if path.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and hoja != "Sheet":
            wb.remove(wb["Sheet"])
        ws = wb[hoja] if hoja in wb.sheetnames else wb.create_sheet(hoja)
        for i, fila in enumerate(filas):
            for j, v in enumerate(fila, 1):
                ws.cell(row=fila_inicio + i, column=j, value=v)
        wb.save(path)
        state_store.log("contabilidad", f"excel-rango:{hoja}", f"{len(filas)} filas")
        return "excel"

    def _filas_demanda(self, mensual: pd.DataFrame, escenario: str) -> list[list]:
        filas = [self.ENC_DEMANDA]
        for i, (_, r) in enumerate(mensual.reset_index(drop=True).iterrows(), 1):
            filas.append([escenario, i, r["etiqueta"],
                          int(r["P1-CC350-RGB_unidades"]),
                          int(r["P2-QT1500-PET_unidades"]),
                          int(r["P3-GARR25L_unidades"])])
        return filas

    def publicar_demanda(self, mensual: pd.DataFrame, escenario: str = "Base") -> str:
        """Demanda BASE -> hoja 'Demanda' (alimenta la hoja Financiero)."""
        return self._escribir_rango("Demanda", self._filas_demanda(mensual, escenario))

    ENC_INVENTARIOS = ["ts", "escenario", "sku", "punto_reorden_s",
                       "stock_seguridad", "lote_Q", "pallets_por_lote",
                       "fill_rate", "capital_inmovilizado_cop"]

    def publicar_inventarios(self, politicas: list[dict]) -> str:
        """Politica (s,Q) del ERP -> hoja 'Inventarios' (rango fijo A4:I8).
        `politicas`: filas de la tabla inventario_politicas (una por SKU)."""
        filas = [self.ENC_INVENTARIOS]
        for p in politicas[:4]:
            filas.append([p.get("ts", ""), p.get("escenario", ""), p.get("sku", ""),
                          p.get("punto_reorden_s", ""), p.get("stock_seguridad", ""),
                          p.get("lote_Q", ""), p.get("pallets_por_lote", ""),
                          p.get("fill_rate", ""), p.get("capital_inmovilizado_cop", "")])
        return self._escribir_rango("Inventarios", filas)

    def publicar_demanda_escenario(self, mensual: pd.DataFrame,
                                   escenario: str) -> str:
        """Demanda del ESCENARIO ACTIVO del ERP -> hoja 'DemandaEscenario'
        (alimenta la hoja FinancieroEscenario para evaluar el escenario)."""
        return self._escribir_rango("DemandaEscenario",
                                    self._filas_demanda(mensual, escenario))

    def leer_parametros(self) -> dict:
        """Lee la hoja Parametros (pares clave-valor) del libro en Drive; asi el
        libro de Sheets puede gobernar parametros del ERP. Fallback: Excel local.

        Contrato de claves que hoy consume `core.finanzas_negocio` (todas
        opcionales; si faltan, el motor usa su default local): TRM,
        GBP_COP, FACTOR_RFQ, TMAR_ANUAL, UPLIFT_THROUGHPUT, FACTOR_MONETIZACION,
        RAMPA_MES5, SCRAP_PP, MANT_EVITADO_MES, AHORRO_LABORAL_MES,
        TASA_RENTA, WC_PCT_INGRESO,
        CRECIMIENTO_DEMANDA_ANUAL, FASES_CAPEX ("0.20,0.35,0.27,0.18"),
        NOMINA_OPERACION_MES, NOMINA_IMPLEMENTACION_MES, OTROS_FIJOS_BASE_MES,
        OTROS_FIJOS_PROYECTO_MES, OPEX_LICENCIAS_MES, CAPEX_SOFTWARE,
        CONTINGENCIA, VIDA_equipos/VIDA_automatizacion/VIDA_servicios/
        VIDA_intangibles/VIDA_software (anios), y unit economics por SKU:
        precio_venta_cop_<SKU> / costo_material_cop_<SKU> (p.ej.
        `precio_venta_cop_P1-CC350-RGB`)."""
        try:
            if self.modo == "sheets":
                ws = self._spreadsheet().worksheet("Parametros")
                # No usar get_all_values() aqui: devuelve el valor VISUAL y
                # puede redondear entradas (p.ej. TRM 3248.87 se muestra
                # "3.249" y Python terminaria usando 3249). El valor sin
                # formato conserva la precision que usan las formulas de
                # Sheets y mantiene ambos motores reconciliados.
                filas = ws.get("A1:D200", value_render_option="UNFORMATTED_VALUE")
            else:
                raise RuntimeError("modo excel")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            from openpyxl import load_workbook
            if not settings.LEDGER_XLSX.exists():
                return {}
            wb = load_workbook(settings.LEDGER_XLSX, read_only=True)
            if "Parametros" not in wb.sheetnames:
                return {}
            filas = [[c if c is not None else "" for c in row]
                     for row in wb["Parametros"].iter_rows(values_only=True)]
        out = {}
        for fila in filas:
            if len(fila) >= 2 and str(fila[0]).strip():
                out[str(fila[0]).strip()] = fila[1]
        return out

    # alias de encabezado -> campo canonico (minusculas, sin acentos ni puntos)
    # tolera variantes reales del libro (p.ej. "activo / paquete", "vida (años)",
    # "cant.") y columnas extra intercaladas (p.ej. "CAPEX COP" ya calculado)
    _ALIAS_CAPEX = {
        "seccion": "seccion", "linea": "linea",
        "activo": "activo", "activo / paquete": "activo", "activo/paquete": "activo",
        "cant": "cantidad", "cant.": "cantidad", "cantidad": "cantidad",
        "moneda": "moneda",
        "costo unitario": "costo_unitario", "costo_unitario": "costo_unitario",
        "vida (años)": "vida_anios", "vida (anos)": "vida_anios",
        "vida_anios": "vida_anios", "vida anios": "vida_anios",
        "categoria d&a": "categoria_dep", "categoria dep": "categoria_dep",
        "categoria_dep": "categoria_dep",
    }
    _CAMPOS_CAPEX = ["seccion", "linea", "activo", "cantidad", "moneda",
                     "costo_unitario", "vida_anios", "categoria_dep"]

    def leer_capex(self) -> list[tuple]:
        """Lee la hoja 'CAPEX' (tabla; mismo esquema logico que CAPEX_FILAS de
        core.finanzas_negocio: seccion, linea, activo, cantidad, moneda,
        costo_unitario, vida_anios, categoria_dep) para que el CAPEX se
        gobierne desde Sheets en vez de la constante local.

        Busca la fila de encabezado por NOMBRE de columna (via `_ALIAS_CAPEX`,
        tolera variantes de redaccion y columnas extra intercaladas como una
        'CAPEX COP' ya calculada) en vez de exigir una lista exacta y
        posicional — el libro real trae textos como 'activo / paquete' o
        'vida (años)' que no calzan letra por letra con el nombre de campo.
        Fallback: lista vacia si la hoja no existe, esta vacia o no se
        encuentra un encabezado reconocible — el motor cae a su CAPEX_FILAS
        local."""
        try:
            if self.modo == "sheets":
                ws = self._spreadsheet().worksheet("CAPEX")
                filas = ws.get("A1:I500", value_render_option="UNFORMATTED_VALUE")
            else:
                raise RuntimeError("modo excel")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            from openpyxl import load_workbook
            if not settings.LEDGER_XLSX.exists():
                return []
            wb = load_workbook(settings.LEDGER_XLSX, read_only=True)
            if "CAPEX" not in wb.sheetnames:
                return []
            filas = [[c if c is not None else "" for c in row]
                     for row in wb["CAPEX"].iter_rows(values_only=True)]

        indice_col: dict[str, int] | None = None
        fila_encabezado = -1
        for i, fila in enumerate(filas):
            candidato = {self._ALIAS_CAPEX[str(c).strip().lower()]: j
                        for j, c in enumerate(fila)
                        if str(c).strip().lower() in self._ALIAS_CAPEX}
            if set(self._CAMPOS_CAPEX) <= set(candidato):
                indice_col, fila_encabezado = candidato, i
                break
        if indice_col is None:
            return []

        salida = []
        for fila in filas[fila_encabezado + 1:]:
            if len(fila) <= max(indice_col.values()) or not str(fila[indice_col["seccion"]]).strip():
                continue
            valores: dict[str, object] = {}
            for c in self._CAMPOS_CAPEX:
                crudo = fila[indice_col[c]]
                if c in ("cantidad", "costo_unitario", "vida_anios"):
                    valores[c] = numero_cop(crudo)
                else:
                    valores[c] = str(crudo).strip()
            if any(valores[c] is None for c in ("cantidad", "costo_unitario", "vida_anios")):
                continue  # fila mal diligenciada (o de resumen/total): se ignora
            salida.append(tuple(valores[c] for c in self._CAMPOS_CAPEX))
        return salida

    def leer_licencias(self) -> dict:
        """Lee de la hoja 'Licencias' los totales 'CAPEX software
        capitalizable' y 'OPEX mensual licencias' (ultima celda no vacia de
        esa fila -- el libro real no los deja en una columna fija) para que
        core.finanzas_negocio los use como override de CAPEX_SOFTWARE /
        OPEX_LICENCIAS_MES. Fallback: {} si Sheets no esta disponible o la
        hoja no tiene esas etiquetas."""
        try:
            if self.modo == "sheets":
                ws = self._spreadsheet().worksheet("Licencias")
                filas = ws.get("A1:I200", value_render_option="UNFORMATTED_VALUE")
            else:
                raise RuntimeError("modo excel")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            from openpyxl import load_workbook
            if not settings.LEDGER_XLSX.exists():
                return {}
            wb = load_workbook(settings.LEDGER_XLSX, read_only=True)
            if "Licencias" not in wb.sheetnames:
                return {}
            filas = [[c if c is not None else "" for c in row]
                     for row in wb["Licencias"].iter_rows(values_only=True)]
        out: dict[str, object] = {}
        for fila in filas:
            if not fila or not str(fila[0]).strip():
                continue
            etiqueta = str(fila[0]).strip().lower()
            # Preservar numeros como int/float: al convertir un valor
            # UNFORMATTED como 8262150.05 a texto, numero_cop interpretaria
            # el punto segun el locale colombiano. _num() acepta el numero
            # crudo directamente y conserva sus decimales.
            no_vacias = [c for c in fila[1:] if str(c).strip()]
            if not no_vacias:
                continue
            if etiqueta.startswith("capex software capitalizable"):
                out["CAPEX_SOFTWARE"] = no_vacias[-1]
            elif etiqueta.startswith("opex mensual licencias"):
                out["OPEX_LICENCIAS_MES"] = no_vacias[-1]
        return out

    def leer_apu_ingenieria(self) -> dict:
        """Lee la hoja 'APU_Ingenieria' (Analisis de Precios Unitarios de los
        costos de ingenieria que cobra ULogix: ingenieria de detalle/FAT/SAT/
        PMO, instalacion/EPC, capacitacion/gestion del cambio — ver
        `tools/publicar_apu_ingenieria.py`). El motor Python no la consume
        directamente: el publicador enlaza sus totales con el costo unitario
        de las 3 filas `Servicios` de `CAPEX`, y el motor lee luego `CAPEX`.
        Devuelve `{'resumen': [...], 'detalle': [...]}` (listas de
        dict), vacias si la hoja no existe o no tiene el formato esperado —
        la página *Finanzas* oculta la sección en ese caso, no revienta."""
        try:
            if self.modo == "sheets":
                ws = self._spreadsheet().worksheet("APU_Ingenieria")
                filas = ws.get_all_values()
            else:
                raise RuntimeError("modo excel")
        except Exception as e:  # noqa: BLE001
            self._sin_fallback(e)
            from openpyxl import load_workbook
            if not settings.LEDGER_XLSX.exists():
                return {"resumen": [], "detalle": []}
            wb = load_workbook(settings.LEDGER_XLSX, read_only=True)
            if "APU_Ingenieria" not in wb.sheetnames:
                return {"resumen": [], "detalle": []}
            filas = [[c if c is not None else "" for c in row]
                     for row in wb["APU_Ingenieria"].iter_rows(values_only=True)]

        def _fila_con(etiqueta: str) -> int | None:
            return next((i for i, f in enumerate(filas) if f and str(f[0]).strip() == etiqueta),
                       None)

        _NUM = {"costo_directo_cop", "pct_administracion", "pct_imprevistos", "pct_utilidad",
               "pct_aiu_total", "aiu_cop", "precio_total_cop", "cantidad", "valor_unitario_cop",
               "subtotal_cop"}

        def _fila_a_dict(encabezado: list[str], fila: list[str]) -> dict:
            d = dict(zip(encabezado, fila))
            for k in list(d):
                if k in _NUM:
                    d[k] = numero_cop(d[k], d[k])
            return d

        resumen: list[dict] = []
        i = _fila_con("RESUMEN")
        if i is not None and i + 1 < len(filas):
            enc = [str(c).strip() for c in filas[i + 1]]
            for f in filas[i + 2:]:
                if not any(str(c).strip() for c in f):
                    break
                resumen.append(_fila_a_dict(enc, f))

        detalle: list[dict] = []
        i = _fila_con("DETALLE")
        if i is not None and i + 1 < len(filas):
            enc = [str(c).strip() for c in filas[i + 1]]
            for f in filas[i + 2:]:
                if any(str(c).strip() for c in f):
                    detalle.append(_fila_a_dict(enc, f))
        return {"resumen": resumen, "detalle": detalle}

    def leer_proveedores_capex(self) -> pd.DataFrame:
        """Registro de proveedores/cotizaciones que acompaña el CAPEX vivo."""
        try:
            if self.modo != "sheets":
                return pd.DataFrame()
            filas = self._spreadsheet().worksheet("Proveedores_CAPEX").get_all_values()
        except Exception:
            return pd.DataFrame()
        idx = next((i for i, f in enumerate(filas)
                    if f and str(f[0]).strip().lower() == "linea"), None)
        if idx is None:
            return pd.DataFrame()
        enc = filas[idx]
        datos = [f + [""] * (len(enc) - len(f)) for f in filas[idx + 1:]
                 if any(str(c).strip() for c in f)]
        return pd.DataFrame(datos, columns=enc)

    def leer_viabilidad_automatizacion(self) -> pd.DataFrame:
        """Resumen antes/Ulogix/mercado publicado en el libro financiero."""
        try:
            if self.modo != "sheets":
                return pd.DataFrame()
            filas = self._spreadsheet().worksheet("Viabilidad_Automatizacion").get_all_values()
        except Exception:
            return pd.DataFrame()
        idx = next((i for i, f in enumerate(filas)
                    if f and str(f[0]).strip() == "Concepto"), None)
        if idx is None:
            return pd.DataFrame()
        enc = filas[idx]
        datos = [f + [""] * (len(enc) - len(f)) for f in filas[idx + 1:]
                 if f and str(f[0]).strip()]
        return pd.DataFrame(datos, columns=enc)

    def probar(self) -> dict:
        """Prueba de conectividad: escribe y relee una celda de verificacion."""
        from datetime import datetime, timezone
        marca = datetime.now(timezone.utc).isoformat(timespec="seconds")
        try:
            destino = self._escribir("_PruebaAPI", ["ts", "origen"],
                                     [[marca, "ulogix-pruebas"]])
            if destino == "sheets":
                ws = self._spreadsheet().worksheet("_PruebaAPI")
                ultimo = ws.get_all_values()[-1][0]
                tabs = [w.title for w in self._spreadsheet().worksheets()]
                return {"ok": ultimo == marca, "modo": "sheets",
                        "detalle": f"celda escrita y releida OK; hojas: {tabs}"}
            return {"ok": True, "modo": "excel",
                    "detalle": f"sin credenciales Sheets: escrito en {settings.LEDGER_XLSX.name}"}
        except Exception as e:  # noqa: BLE001
            return {"ok": False, "modo": self.modo, "detalle": str(e)}


def libro_local() -> pd.DataFrame:
    """Lee el libro de produccion local (respaldo Excel) si existe."""
    if settings.LEDGER_XLSX.exists():
        try:
            return pd.read_excel(settings.LEDGER_XLSX, sheet_name="LibroProduccion")
        except ValueError:
            pass
    return pd.DataFrame(columns=ENCABEZADOS_LIBRO)
